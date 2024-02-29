from pdf2docx import Converter
import argparse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image

parser= argparse.ArgumentParser(
        description="This is a converter in which you can convert pdf file to ppt,excel,docx.",
        usage='%(prog)s -ppt -docx -excel ' ,
        epilog="example : %(prog)s -ppt file_name.pdf",           
        )
parser.add_argument(
        "-ppt",
        help="uplode pdf file",
        metavar="pdf to ppt file",
        dest="ppt",
        nargs="+" 
        )
parser.add_argument(
        "-docx",
        help="uplode pdf  file",
        metavar="pdf to docx file",
        dest="docx",
        nargs="+" 
        )
parser.add_argument(
        "-excel",
        help="uplode pdf file",
        metavar="pdf to excel file ",
        dest="excel",
        nargs="+" 
        )
args=parser.parse_args()

pdf1=args.ppt
pdf2=args.docx
pdf3=args.excel
if pdf1:
        ppt='newresume.ppt'
        cv=Converter(pdf1[0])
        cv.convert(ppt)
elif pdf2:    
        docx='newresume.docx'
        cv=Converter(pdf2[0])
        cv.convert(docx)
        
# elif pdf3:
#         # excel='newsynopsis.xlsx'
#         # cv=Converter(pdf3[0])
#         # cv.convert(excel)
        
#         im =pdf3[0]
#         print(im)
#         # im =Image.open(pdf3[0])
#         # print(im)
#         rgb_im = im.convert('RGB')
        
#         rgbArr = []
        
#         def rgbToHex(rgb):
#                 return '%02x%02x%02x' % rgb

#         def rgbArray():
#                 imgX = im.size[0]
#                 imgY = im.size[1]

#                 for i in range(imgY):
#                         rgbArr.append([])

#                 for i in range(imgX):
#                         for j in range(imgY):
#                                 r, g, b = rgb_im.getpixel((i, j))
                        
#                         rgbArr[j].append([r, g, b])
                
#                 return rgbArr
#         # print(rgbArr)
#         def generateSheet():
#                 wb = Workbook()

#                 ws =  wb.active
#                 ws.title = "image"

#                 for i in range(im.size[0]):
#                         ws.column_dimensions[get_column_letter(i + 1)].width = 3.0
#                         for j in range(im.size[1]):
#                                 cell = get_column_letter(i + 1) + str(j + 1)
#                                 ws[cell].fill = PatternFill(fgColor = rgbArr[j][i], fill_type = "solid")

#                 wb.save(filename = 'image.xlsx')

#                 rgbArray()

#         for i in range(len(rgbArr)):
#                 for j in range(len(rgbArr[i])):
#                         rgbArr[i][j] = rgbToHex((rgbArr[i][j][0], rgbArr[i][j][1], rgbArr[i][j][2]))

#                 print('this tool currently only works with image files, large files may also take long to process/not work entirely')
#                 generateSheet()

#                 print('\ndone')
